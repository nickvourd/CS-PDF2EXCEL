#!/usr/bin/env python3

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BANNER = r"""
 ██████╗███████╗      ██████╗ ██████╗ ███████╗██████╗ ███████╗██╗  ██╗ ██████╗███████╗██╗
██╔════╝██╔════╝      ██╔══██╗██╔══██╗██╔════╝╚════██╗██╔════╝╚██╗██╔╝██╔════╝██╔════╝██║
██║     ███████╗█████╗██████╔╝██║  ██║█████╗   █████╔╝█████╗   ╚███╔╝ ██║     █████╗  ██║
██║     ╚════██║╚════╝██╔═══╝ ██║  ██║██╔══╝  ██╔═══╝ ██╔══╝   ██╔██╗ ██║     ██╔══╝  ██║
╚██████╗███████║      ██║     ██████╔╝██║     ███████╗███████╗██╔╝ ██╗╚██████╗███████╗███████╗
 ╚═════╝╚══════╝      ╚═╝     ╚═════╝ ╚═╝     ╚══════╝╚══════╝╚═╝  ╚═╝ ╚═════╝╚══════╝╚══════╝

                          Created with <3 by @nickvourd
"""

DATE_RE = re.compile(r"^\d{2}/\d{2} \d{2}:\d{2}\b")
META_RE = re.compile(r"^(User|Process|PID|Opened):\s*(.*)$")
PAGE_FOOTER_RE = re.compile(r"^Page\.\s*\d+\s*$")
PAGE_HEADER = "Sessions Report"

SECTION_HEADERS = {"Communication Path", "File Hashes", "Activity"}
TABLE_HEADERS = {
    "hosts port protocol",
    "date hash name",
    "date activity",
}


@dataclass
class Session:
    session_id: int
    hostname: str
    user: str = ""
    process: str = ""
    pid: str = ""
    opened: str = ""
    comm_paths: list[tuple[str, str, str]] = field(default_factory=list)
    file_hashes: list[tuple[str, str, str]] = field(default_factory=list)
    activities: list[tuple[str, str]] = field(default_factory=list)


def extract_text_lines(pdf_path: Path) -> list[str]:
    lines: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[1:]:
            text = page.extract_text() or ""
            for raw in text.splitlines():
                line = raw.rstrip()
                if not line.strip():
                    continue
                if line.strip() == PAGE_HEADER:
                    continue
                if PAGE_FOOTER_RE.match(line.strip()):
                    continue
                lines.append(line)
    return lines


_HOSTNAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]*$")


def _is_session_header(line: str, next_lines: Iterable[str]) -> bool:
    stripped = line.strip()
    if not stripped or " " in stripped:
        return False
    if stripped in SECTION_HEADERS or stripped in TABLE_HEADERS:
        return False
    if META_RE.match(stripped):
        return False
    if not _HOSTNAME_RE.match(stripped):
        return False
    peek = next(iter(next_lines), "")
    return peek.strip().startswith("User:")


def parse_sessions(lines: list[str]) -> list[Session]:
    sessions: list[Session] = []
    current: Session | None = None
    section: str | None = None  # "comm" | "hashes" | "activity" | None
    pending_row: list[str] | None = None  # accumulator for multi-line table rows

    def flush_pending() -> None:
        nonlocal pending_row
        if pending_row is None or current is None:
            pending_row = None
            return
        if section == "activity":
            date, *rest = pending_row
            current.activities.append((date, " ".join(rest).strip()))
        elif section == "hashes":
            if len(pending_row) >= 3:
                date, hashv, *name_parts = pending_row
                name = " ".join(name_parts).strip()
                name = re.sub(r"\s*\\\s*", r"\\", name)
                current.file_hashes.append((date, hashv, name))
            elif len(pending_row) == 2:
                current.file_hashes.append((pending_row[0], pending_row[1], ""))
        elif section == "comm":
            # Expected: hosts, port, protocol
            if len(pending_row) >= 3:
                current.comm_paths.append(
                    (pending_row[0], pending_row[1], " ".join(pending_row[2:]))
                )
            elif len(pending_row) == 2:
                current.comm_paths.append((pending_row[0], pending_row[1], ""))
        pending_row = None

    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        next_lines = lines[i + 1 : i + 4]

        # Section transitions
        if stripped == "Communication Path":
            flush_pending()
            section = "comm"
            i += 1
            continue
        if stripped == "File Hashes":
            flush_pending()
            section = "hashes"
            i += 1
            continue
        if stripped == "Activity":
            flush_pending()
            section = "activity"
            i += 1
            continue

        # Skip table header rows
        if stripped in TABLE_HEADERS:
            flush_pending()
            i += 1
            continue

        # New session block
        if _is_session_header(line, next_lines):
            flush_pending()
            session = Session(session_id=len(sessions) + 1, hostname=stripped)
            sessions.append(session)
            current = session
            section = None
            i += 1
            continue

        # Metadata fields
        meta = META_RE.match(stripped)
        if meta and current is not None and section is None:
            key, value = meta.group(1).lower(), meta.group(2).strip()
            setattr(current, key, value)
            i += 1
            continue

        # Table content rows
        if section in {"activity", "hashes", "comm"} and current is not None:
            if DATE_RE.match(stripped) or (
                section == "comm" and pending_row is None
            ):
                # Row starts here -> flush previous accumulation, begin new row
                flush_pending()
                if section == "comm":
                    parts = stripped.rsplit(maxsplit=2)
                    pending_row = parts
                    flush_pending()
                else:
                    date = stripped[:11]  # "MM/DD HH:MM"
                    remainder = stripped[11:].strip()
                    if section == "hashes":
                        hash_match = re.match(r"(\S+)\s*(.*)", remainder)
                        if hash_match:
                            pending_row = [date, hash_match.group(1), hash_match.group(2)]
                        else:
                            pending_row = [date, "", ""]
                    else:
                        pending_row = [date, remainder]
            else:
                # Continuation of the previous row (wrapped text)
                if pending_row is not None:
                    pending_row.append(stripped)
            i += 1
            continue

        # Anything else -> ignore
        i += 1

    flush_pending()
    return sessions


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="305496")
BODY_FONT = Font(name="Arial")
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = BODY_FONT
            cell.alignment = WRAP

    # Reasonable column widths
    widths = {
        "session_id": 11,
        "hostname": 16,
        "user": 14,
        "process": 36,
        "pid": 8,
        "opened": 14,
        "host": 36,
        "port": 8,
        "protocol": 10,
        "date": 14,
        "hash": 36,
        "name": 60,
        "activity": 70,
    }
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 18)
    ws.freeze_panes = "A2"


def write_excel(sessions: list[Session], out_path: Path) -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "Sessions",
        ["session_id", "hostname", "user", "process", "pid", "opened"],
        [
            [s.session_id, s.hostname, s.user, s.process, s.pid, s.opened]
            for s in sessions
        ],
    )
    _write_sheet(
        wb,
        "CommunicationPaths",
        ["session_id", "hostname", "host", "port", "protocol"],
        [
            [s.session_id, s.hostname, host, port, proto]
            for s in sessions
            for host, port, proto in s.comm_paths
        ],
    )
    _write_sheet(
        wb,
        "FileHashes",
        ["session_id", "hostname", "date", "hash", "name"],
        [
            [s.session_id, s.hostname, date, hashv, name]
            for s in sessions
            for date, hashv, name in s.file_hashes
        ],
    )
    _write_sheet(
        wb,
        "Activities",
        ["session_id", "hostname", "date", "activity"],
        [
            [s.session_id, s.hostname, date, activity]
            for s in sessions
            for date, activity in s.activities
        ],
    )

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description=BANNER + "\n" + (__doc__ or ""),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("pdf", type=Path, help="Path to the Sessions Report PDF")
    ap.add_argument(
        "xlsx",
        type=Path,
        nargs="?",
        help="Output Excel path (defaults to <pdf>.xlsx next to the input)",
    )
    args = ap.parse_args()

    print(BANNER)

    if not args.pdf.is_file():
        print(f"Input not found: {args.pdf}", file=sys.stderr)
        return 2

    out = args.xlsx or args.pdf.with_suffix(".xlsx")
    lines = extract_text_lines(args.pdf)
    sessions = parse_sessions(lines)
    write_excel(sessions, out)

    total_acts = sum(len(s.activities) for s in sessions)
    total_hashes = sum(len(s.file_hashes) for s in sessions)
    total_comms = sum(len(s.comm_paths) for s in sessions)
    print(
        f"Wrote {out}\n"
        f"  Sessions:            {len(sessions)}\n"
        f"  Communication paths: {total_comms}\n"
        f"  File hashes:         {total_hashes}\n"
        f"  Activities:          {total_acts}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
